from flask import ( 
    Blueprint, render_template,
      request, flash, redirect,
        url_for, abort, current_app,
          send_file )
from flask_login import login_required, current_user
from io import BytesIO
import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from decimal import Decimal
from . import db
from .models import Device, User, Customer, Category, Product, Supplier  # add User if not imported
from flask import abort
import os
from werkzeug.utils import secure_filename
import uuid
import xlrd
import pandas as pd
from sqlalchemy import or_
from datetime import datetime
from .models import User, Product, Customer, Outgoing, Purchase
import io
from werkzeug.security import generate_password_hash
from website.models import User, Product, Customer, Supplier, Device

from functools import wraps

def roles_required(*roles):
    """Decorator: require current_user.role to be in roles.
    Falls back to attribute `is_admin` if `role` is not present.
    Usage: @roles_required('admin') above a view function."""
    def decorator(f):
        @wraps(f)
        def wrapped(*args, **kwargs):
            # ensure user is logged in
            if not current_user or not current_user.is_authenticated:
                abort(403)

            # prefer string role if model has it
            user_role = getattr(current_user, "role", None)
            if user_role is None:
                # fallback to boolean is_admin property (older model)
                is_admin_flag = getattr(current_user, "is_admin", False)
                # if 'admin' is required and user has is_admin True -> allow
                if "admin" in roles and is_admin_flag:
                    return f(*args, **kwargs)
                # otherwise not allowed
                abort(403)

            # if role exists, check it is allowed
            if user_role not in roles:
                abort(403)

            return f(*args, **kwargs)
        return wrapped
    return decorator

views = Blueprint('views', __name__)


@views.route('/', methods=['GET', 'POST'])
@login_required
def home():
    # only admins should see the admin dashboard (optional)
    # if not current_user.is_admin:
    #     return render_template("home.html", user=current_user)

    stats = {
        "system_users": User.query.count(),
        "categories": Category.query.count(),
        "products": Product.query.count(),
        "customers": Customer.query.count(),
        "suppliers": Supplier.query.count(),
        # if you want number of outgoing records; change to .with_entities(db.func.sum(Outgoing.qty)) if you want total qty
        "total_outgoing": Outgoing.query.count(),
    }

    return render_template(
        "admin_home.html",    # your dashboard template filename
        user=current_user,
        stats=stats,
    )


@views.route("/admin/categories", methods=["GET", "POST"])
@login_required
def category_list():
    if not current_user.is_admin:
        abort(403)

    # --- Handle ADD new category (POST) ---
    if request.method == "POST":
        name = (request.form.get("name") or "").strip()
        if not name:
            flash("Category name is required.", "error")
            return redirect(url_for("views.category_list"))

        existing = Category.query.filter_by(name=name).first()
        if existing:
            flash("Category name already exists.", "error")
            return redirect(url_for("views.category_list"))

        new_cat = Category(name=name)
        db.session.add(new_cat)
        db.session.commit()
        flash("Category added successfully.", "success")
        return redirect(url_for("views.category_list"))

    # --- Handle LIST (GET) with search + pagination ---
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 10, type=int)
    search = (request.args.get("q") or "").strip()

    query = Category.query

    if search:
        query = query.filter(Category.name.ilike(f"%{search}%"))

    pagination = query.order_by(Category.id).paginate(
        page=page,
        per_page=per_page,
        error_out=False,
    )

    categories = pagination.items

    return render_template(
        "admin_categories.html",
        user=current_user,
        categories=categories,
        pagination=pagination,
        per_page=per_page,
        search=search,
    )


@views.route("/admin/categories/<int:category_id>/delete", methods=["POST"])
@login_required
def delete_category(category_id):
    if not current_user.is_admin:
        abort(403)

    category = Category.query.get_or_404(category_id)
    db.session.delete(category)
    db.session.commit()
    flash("Category deleted.", "success")
    return redirect(url_for("views.category_list"))

@views.route("/admin/products", methods=["GET"])
@login_required
def product_list():
    if not current_user.is_admin:
        abort(403)

    # Query params
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 10, type=int)
    search = request.args.get("q", "", type=str).strip()
    category_id = request.args.get("category_id", type=int)

    query = Product.query

    # Apply search (by name)
    if search:
        query = query.filter(Product.name.ilike(f"%{search}%"))

    # Apply category filter
    if category_id:
        query = query.filter(Product.category_id == category_id)

    # Pagination
    pagination = query.order_by(Product.id).paginate(
        page=page,
        per_page=per_page,
        error_out=False,
    )
    products = pagination.items

    # For filter dropdown + edit modal
    categories = Category.query.order_by(Category.name).all()

    return render_template(
        "admin_products.html",
        user=current_user,
        products=products,
        categories=categories,
        pagination=pagination,
        search=search,
        selected_category=category_id,
        per_page=per_page,
    )


@views.route("/admin/products/new", methods=["GET", "POST"])
@login_required
def product_create():
    if not current_user.is_admin:
        abort(403)

    categories = Category.query.order_by(Category.name).all()

    if request.method == "POST":
        name = (request.form.get("name") or "").strip()
        price_raw = (request.form.get("price") or "").strip()
        qty_raw = (request.form.get("quantity") or "").strip()
        category_id = request.form.get("category_id") or None

        errors = []

        # --- VALIDATION SECTION ---
        if not name:
            errors.append("Product name is required.")

        # price validation
        price = None
        if not price_raw:
            errors.append("Price is required.")
        else:
            try:
                price = Decimal(price_raw)
                if price < 0:
                    errors.append("Price cannot be negative.")
            except Exception:
                errors.append("Price must be a number.")

        # quantity validation
        quantity = 0
        if not qty_raw:
            errors.append("Quantity is required.")
        else:
            try:
                quantity = int(qty_raw)
                if quantity < 0:
                    errors.append("Quantity cannot be negative.")
            except ValueError:
                errors.append("Quantity must be an integer.")

        # category validation
        category_obj = None
        if category_id:
            category_obj = Category.query.get(category_id)
            if not category_obj:
                errors.append("Selected category does not exist.")
        else:
            errors.append("Category is required.")

        # If errors, re-render form
        if errors:
            for e in errors:
                flash(e, "error")
            return render_template(
                "admin_product_form.html",
                user=current_user,
                categories=categories,
                form_data={
                    "name": name,
                    "price": price_raw,
                    "quantity": qty_raw,
                    "category_id": category_id,
                },
            )

        # ---------------------------------------------
        # 🔥 IMAGE UPLOAD HANDLING (ADD THIS BLOCK)
        # ---------------------------------------------
        file = request.files.get("image_file")
        filename = None

        if file and file.filename != "":
            if current_app.allowed_file(file.filename):
                original = secure_filename(file.filename)
                ext = original.rsplit(".", 1)[1].lower()

                # unique filename
                filename = f"{uuid.uuid4().hex}.{ext}"

                save_path = os.path.join(current_app.config["UPLOAD_FOLDER"], filename)
                file.save(save_path)
            else:
                flash("Invalid image type. Allowed: png, jpg, jpeg, gif", "error")
                return redirect(request.url)

        # ---------------------------------------------
        # 🔥 CREATE PRODUCT IN DATABASE
        # ---------------------------------------------
        new_product = Product(
            name=name,
            price=price,
            quantity=quantity,
            category_id=category_obj.id,
            image_filename=filename,   # << store uploaded file
        )

        db.session.add(new_product)
        db.session.commit()

        flash("Product created successfully.", "success")
        return redirect(url_for("views.product_list"))

    # GET: show form
    return render_template(
        "admin_product_form.html",
        user=current_user,
        categories=categories,
        form_data=None,
    )

@views.route("/admin/products/<int:product_id>/edit", methods=["POST"])
@login_required
def product_edit(product_id):
    if not current_user.is_admin:
        abort(403)

    product = Product.query.get_or_404(product_id)
    categories = Category.query.order_by(Category.name).all()

    # Form Data
    name = (request.form.get("name") or "").strip()
    price_raw = (request.form.get("price") or "").strip()
    qty_raw = (request.form.get("quantity") or "").strip()
    category_id = request.form.get("category_id") or None

    # Validation
    errors = []

    if not name:
        errors.append("Product name is required.")

    # Price validation
    try:
        price = Decimal(price_raw)
        if price < 0:
            errors.append("Price cannot be negative.")
    except:
        errors.append("Invalid price format.")

    # Quantity validation
    try:
        quantity = int(qty_raw)
        if quantity < 0:
            errors.append("Quantity cannot be negative.")
    except:
        errors.append("Invalid quantity (must be an integer).")

    # Category validation
    category = Category.query.get(category_id)
    if not category:
        errors.append("Selected category does not exist.")

    # Handle file upload
    file = request.files.get("image_file")
    filename = product.image_filename  # keep old image if none uploaded

    if file and file.filename != "":
        if current_app.allowed_file(file.filename):
            filename = secure_filename(file.filename)
            save_path = os.path.join(current_app.config["UPLOAD_FOLDER"], filename)
            file.save(save_path)
        else:
            flash("Invalid image file type!", "error")
            return redirect(url_for("views.product_list"))

    # If errors → show flash messages
    if errors:
        for e in errors:
            flash(e, "error")
        return redirect(url_for("views.product_list"))

    # Update product
    product.name = name
    product.price = price
    product.quantity = quantity
    product.category_id = category.id
    product.image_filename = filename

    db.session.commit()
    flash("Product updated successfully!", "success")
    return redirect(url_for("views.product_list"))


@views.route("/admin/products/<int:product_id>/delete", methods=["POST"])
@login_required
def product_delete(product_id):
    if not current_user.is_admin:
        abort(403)

    product = Product.query.get_or_404(product_id)
    db.session.delete(product)
    db.session.commit()
    flash("Product deleted.", "success")
    return redirect(url_for("views.product_list"))

@views.route("/admin/products/import", methods=["POST"])
@login_required
def product_import():
    if not current_user.is_admin:
        abort(403)

    file = request.files.get("file")

    if not file or file.filename == "":
        flash("Please choose an Excel file to upload.", "error")
        return redirect(url_for("views.product_list"))

    filename = file.filename
    ext = filename.rsplit(".", 1)[-1].lower()

    # Allowed Excel extensions
    allowed_exts = {"xlsx", "xlsm", "xltx", "xltm", "xls"}
    if ext not in allowed_exts:
        flash("Invalid file type. Allowed: .xls, .xlsx Excel files only.", "error")
        return redirect(url_for("views.product_list"))

    created = 0
    updated = 0
    skipped = 0

    try:
        rows = None

        # ---- openpyxl for xlsx-like formats ----
        if ext in {"xlsx", "xlsm", "xltx", "xltm"}:
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))

        # ---- xlrd for old .xls ----
        elif ext == "xls":
            data = file.read()
            book = xlrd.open_workbook(file_contents=data)
            sheet = book.sheet_by_index(0)
            rows = []
            for r in range(sheet.nrows):
                rows.append(sheet.row_values(r))

        if not rows or len(rows) < 2:
            flash("The uploaded file is empty or has no data rows.", "error")
            return redirect(url_for("views.product_list"))

        # ---- header mapping ----
        header_raw = rows[0]
        header = [
            (str(c).strip().lower() if c is not None else "")
            for c in header_raw
        ]

        def find_col(col_name):
            try:
                return header.index(col_name)
            except ValueError:
                return None

        # required
        name_idx = find_col("name")
        price_idx = find_col("price")
        qty_idx = find_col("quantity")
        category_idx = find_col("category")
        # optional
        image_idx = find_col("image")  # or "image_filename" if you prefer; adjust header

        if None in (name_idx, price_idx, qty_idx, category_idx):
            flash(
                "Header row must contain columns: Name, Price, Quantity, Category "
                "(and optional Image).",
                "error",
            )
            return redirect(url_for("views.product_list"))

        def get_cell(row, i):
            if i is None:
                return ""
            if i >= len(row):
                return ""
            val = row[i]
            return "" if val is None else str(val).strip()

        # ---- process data rows ----
        for row in rows[1:]:
            if row is None:
                continue

            name = get_cell(row, name_idx)
            price_raw = get_cell(row, price_idx)
            qty_raw = get_cell(row, qty_idx)
            category_name = get_cell(row, category_idx)
            image_val = get_cell(row, image_idx) if image_idx is not None else ""

            # skip completely empty rows
            if not (name or price_raw or qty_raw or category_name or image_val):
                continue

            # basic validations
            if not name:
                skipped += 1
                continue

            # price
            try:
                price = Decimal(price_raw)
                if price < 0:
                    skipped += 1
                    continue
            except Exception:
                skipped += 1
                continue

            # quantity
            try:
                quantity = int(float(qty_raw))  # handles "10.0" too
                if quantity < 0:
                    skipped += 1
                    continue
            except Exception:
                skipped += 1
                continue

            # category – by name
            if not category_name:
                skipped += 1
                continue

            category = Category.query.filter_by(name=category_name).first()
            if not category:
                # if you want to autocreate categories, replace this with making a new Category
                skipped += 1
                continue

            # existing product? We'll match by name (and category)
            existing = Product.query.filter_by(
                name=name,
                category_id=category.id
            ).first()

            image_filename = image_val or None  # can be empty

            if existing:
                existing.price = price
                existing.quantity = quantity
                existing.category_id = category.id
                if image_filename:
                    existing.image_filename = image_filename
                updated += 1
            else:
                p = Product(
                    name=name,
                    price=price,
                    quantity=quantity,
                    category_id=category.id,
                    image_filename=image_filename,
                )
                db.session.add(p)
                created += 1

        db.session.commit()
        flash(
            f"Product import complete. Created: {created}, Updated: {updated}, Skipped: {skipped}.",
            "success",
        )

    except Exception as e:
        db.session.rollback()
        flash(f"Failed to import products: {e}", "error")

    return redirect(url_for("views.product_list"))


@views.route("/admin/categories/<int:category_id>/edit", methods=["POST"])
@login_required
def edit_category(category_id):
    if not current_user.is_admin:
        abort(403)

    category = Category.query.get_or_404(category_id)
    new_name = (request.form.get("name") or "").strip()

    if not new_name:
        flash("Category name cannot be empty.", "error")
        return redirect(url_for("views.category_list"))

    category.name = new_name
    db.session.commit()
    flash("Category updated successfully.", "success")
    return redirect(url_for("views.category_list"))


@views.route("/admin/customers", methods=["GET"])
@login_required
def customer_list():

    ## this functions enable even user only
    # if not current_user.is_admin:
    #     abort(403)

    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 10, type=int)
    search = (request.args.get("q") or "").strip()

    query = Customer.query

    if search:
        query = query.filter(Customer.name.ilike(f"%{search}%"))

    pagination = query.order_by(Customer.id).paginate(
        page=page,
        per_page=per_page,
        error_out=False,
    )

    customers = pagination.items

    return render_template(
        "admin_customers.html",
        user=current_user,
        customers=customers,
        pagination=pagination,
        per_page=per_page,
        search=search,
    )

@views.route("/admin/customers/new", methods=["POST"])
@login_required
def customer_create():
    if not current_user.is_admin:
        abort(403)

    name = (request.form.get("name") or "").strip()
    address = (request.form.get("address") or "").strip()
    email = (request.form.get("email") or "").strip()
    contact = (request.form.get("contact") or "").strip()

    errors = []

    if not name:
        errors.append("Name is required.")
    if not address:
        errors.append("Address is required.")
    if not email:
        errors.append("Email is required.")
    if not contact:
        errors.append("Contact is required.")

    if errors:
        for e in errors:
            flash(e, "error")
        return redirect(url_for("views.customer_list"))

    customer = Customer(
        name=name,
        address=address,
        email=email,
        contact=contact,
    )
    db.session.add(customer)
    db.session.commit()
    flash("Customer added successfully.", "success")
    return redirect(url_for("views.customer_list"))
@views.route("/admin/customers/<int:customer_id>/edit", methods=["POST"])
@login_required
def customer_edit(customer_id):
    if not current_user.is_admin:
        abort(403)

    customer = Customer.query.get_or_404(customer_id)

    name = (request.form.get("name") or "").strip()
    address = (request.form.get("address") or "").strip()
    email = (request.form.get("email") or "").strip()
    contact = (request.form.get("contact") or "").strip()

    errors = []

    if not name:
        errors.append("Name is required.")
    if not address:
        errors.append("Address is required.")
    if not email:
        errors.append("Email is required.")
    if not contact:
        errors.append("Contact is required.")

    if errors:
        for e in errors:
            flash(e, "error")
        return redirect(url_for("views.customer_list"))

    customer.name = name
    customer.address = address
    customer.email = email
    customer.contact = contact

    db.session.commit()
    flash("Customer updated successfully.", "success")
    return redirect(url_for("views.customer_list"))

@views.route("/admin/customers/<int:customer_id>/delete", methods=["POST"])
@login_required
def customer_delete(customer_id):
    if not current_user.is_admin:
        abort(403)

    customer = Customer.query.get_or_404(customer_id)

    db.session.delete(customer)
    db.session.commit()
    flash("Customer deleted.", "success")
    return redirect(url_for("views.customer_list"))

@views.route("/admin/customers/export/excel", methods=["GET"])
@login_required
def customer_export_excel():
    if not current_user.is_admin:
        abort(403)

    customers = Customer.query.order_by(Customer.id).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customers"

    headers = ["ID", "Name", "Address", "Email", "Contact"]
    ws.append(headers)

    for c in customers:
        ws.append([c.id, c.name, c.address, c.email, c.contact])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return send_file(
        stream,
        as_attachment=True,
        download_name="customers.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@views.route("/admin/customers/export/pdf", methods=["GET"])
@login_required
def customer_export_pdf():
    if not current_user.is_admin:
        abort(403)

    customers = Customer.query.order_by(Customer.id).all()

    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    y = height - 50
    p.setFont("Helvetica-Bold", 14)
    p.drawString(50, y, "Customer List")
    y -= 30

    p.setFont("Helvetica-Bold", 10)
    p.drawString(50, y, "ID")
    p.drawString(80, y, "Name")
    p.drawString(220, y, "Address")
    p.drawString(400, y, "Email")
    p.drawString(520, y, "Contact")
    y -= 20

    p.setFont("Helvetica", 9)
    for c in customers:
        if y < 50:
            p.showPage()
            y = height - 50
        p.drawString(50, y, str(c.id))
        p.drawString(80, y, c.name[:30])
        p.drawString(220, y, c.address[:35])
        p.drawString(400, y, c.email[:25])
        p.drawString(520, y, c.contact[:15])
        y -= 16

    p.showPage()
    p.save()
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="customers.pdf",
        mimetype="application/pdf",
    )

@views.route("/admin/customers/import", methods=["POST"])
@login_required
def customer_import():
    if not current_user.is_admin:
        abort(403)

    file = request.files.get("file")

    if not file or file.filename == "":
        flash("Please choose an Excel file to upload.", "error")
        return redirect(url_for("views.customer_list"))

    filename = file.filename
    ext = filename.rsplit(".", 1)[-1].lower()

    # Only allow Excel formats
    allowed_exts = {"xlsx", "xlsm", "xltx", "xltm", "xls"}
    if ext not in allowed_exts:
        flash("Invalid file type. Allowed: .xls, .xlsx Excel files only.", "error")
        return redirect(url_for("views.customer_list"))

    created = 0
    updated = 0
    skipped = 0

    try:
        rows = None

        # ---- .xlsx / .xlsm / xltx / xltm via openpyxl ----
        if ext in {"xlsx", "xlsm", "xltx", "xltm"}:
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))

        # ---- legacy .xls via xlrd ----
        elif ext == "xls":
            # xlrd wants bytes
            data = file.read()
            book = xlrd.open_workbook(file_contents=data)
            sheet = book.sheet_by_index(0)
            rows = []
            for r in range(sheet.nrows):
                rows.append(sheet.row_values(r))

        if not rows or len(rows) < 2:
            flash("The uploaded file is empty or has no data rows.", "error")
            return redirect(url_for("views.customer_list"))

        # ---- header mapping ----
        header_raw = rows[0]
        header = [
            (str(c).strip().lower() if c is not None else "")
            for c in header_raw
        ]

        def find_col(col_name):
            try:
                return header.index(col_name)
            except ValueError:
                return None

        name_idx = find_col("name")
        address_idx = find_col("address")
        email_idx = find_col("email")
        contact_idx = find_col("contact")

        if None in (name_idx, address_idx, email_idx, contact_idx):
            flash(
                "Header row must contain columns: Name, Address, Email, Contact.",
                "error",
            )
            return redirect(url_for("views.customer_list"))

        # ---- process data rows ----
        for row in rows[1:]:
            # Some libraries return tuples, others lists – treat the same
            if row is None:
                continue

            # Safely get each column
            def get_cell(i):
                if i is None:
                    return ""
                if i >= len(row):
                    return ""
                val = row[i]
                return "" if val is None else str(val).strip()

            name = get_cell(name_idx)
            address = get_cell(address_idx)
            email = get_cell(email_idx)
            contact = get_cell(contact_idx)

            # Skip completely empty rows
            if not (name or address or email or contact):
                continue

            # Basic validation – must at least have name + email
            if not name or not email:
                skipped += 1
                continue

            # Use email to detect existing customer
            existing = Customer.query.filter_by(email=email).first()

            if existing:
                existing.name = name
                existing.address = address
                existing.contact = contact
                updated += 1
            else:
                c = Customer(
                    name=name,
                    address=address,
                    email=email,
                    contact=contact,
                )
                db.session.add(c)
                created += 1

        db.session.commit()
        msg = f"Import complete. Created: {created}, Updated: {updated}, Skipped: {skipped}."
        flash(msg, "success")

    except Exception as e:
        db.session.rollback()
        flash(f"Failed to import customers: {e}", "error")

    return redirect(url_for("views.customer_list"))


@views.route("/admin/suppliers")
@login_required
def supplier_list():
    if not current_user.is_admin:
        abort(403)

    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 10, type=int)
    search = (request.args.get("q") or "").strip()

    query = Supplier.query

    if search:
        like = f"%{search}%"
        query = query.filter(
            or_(
                Supplier.name.ilike(like),
                Supplier.address.ilike(like),
                Supplier.email.ilike(like),
                Supplier.contact.ilike(like),
            )
        )

    pagination = query.order_by(Supplier.id).paginate(
        page=page, per_page=per_page, error_out=False
    )
    suppliers = pagination.items

    return render_template(
        "admin_suppliers.html",
        user=current_user,
        suppliers=suppliers,
        pagination=pagination,
        search=search,
        per_page=per_page,
    )
@views.route("/admin/suppliers/new", methods=["POST"])
@login_required
def supplier_create():
    if not current_user.is_admin:
        abort(403)

    name = (request.form.get("name") or "").strip()
    address = (request.form.get("address") or "").strip()
    email = (request.form.get("email") or "").strip()
    contact = (request.form.get("contact") or "").strip()

    if not name:
        flash("Supplier name is required.", "error")
        return redirect(url_for("views.supplier_list"))

    supplier = Supplier(
        name=name,
        address=address,
        email=email,
        contact=contact,
    )
    db.session.add(supplier)
    db.session.commit()
    flash("Supplier added successfully.", "success")
    return redirect(url_for("views.supplier_list"))
@views.route("/admin/suppliers/<int:supplier_id>/edit", methods=["POST"])
@login_required
def supplier_edit(supplier_id):
    if not current_user.is_admin:
        abort(403)

    supplier = Supplier.query.get_or_404(supplier_id)

    name = (request.form.get("name") or "").strip()
    address = (request.form.get("address") or "").strip()
    email = (request.form.get("email") or "").strip()
    contact = (request.form.get("contact") or "").strip()

    if not name:
        flash("Supplier name is required.", "error")
        return redirect(url_for("views.supplier_list"))

    supplier.name = name
    supplier.address = address
    supplier.email = email
    supplier.contact = contact

    db.session.commit()
    flash("Supplier updated successfully.", "success")
    return redirect(url_for("views.supplier_list"))
@views.route("/admin/suppliers/<int:supplier_id>/delete", methods=["POST"])
@login_required
def supplier_delete(supplier_id):
    if not current_user.is_admin:
        abort(403)

    supplier = Supplier.query.get_or_404(supplier_id)
    db.session.delete(supplier)
    db.session.commit()
    flash("Supplier deleted.", "success")
    return redirect(url_for("views.supplier_list"))
@views.route("/admin/suppliers/export/excel")
@login_required
def supplier_export_excel():
    if not current_user.is_admin:
        abort(403)

    suppliers = Supplier.query.order_by(Supplier.id).all()

    data = [
        {
            "ID": s.id,
            "Name": s.name,
            "Address": s.address,
            "Email": s.email,
            "Contact": s.contact,
        }
        for s in suppliers
    ]

    df = pd.DataFrame(data)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        df.to_excel(writer, index=False, sheet_name="Suppliers")

    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="suppliers.xlsx",
    )
@views.route("/admin/suppliers/export/pdf")
@login_required
def supplier_export_pdf():
    if not current_user.is_admin:
        abort(403)

    suppliers = Supplier.query.order_by(Supplier.id).all()

    buffer = BytesIO()
    p = canvas.Canvas(buffer)

    y = 800
    p.setFont("Helvetica-Bold", 14)
    p.drawString(50, y, "List of Suppliers")
    y -= 30
    p.setFont("Helvetica", 10)

    for s in suppliers:
        line = f"{s.id}  | {s.name} | {s.address or ''} | {s.email or ''} | {s.contact or ''}"
        p.drawString(50, y, line[:130])  # simple clipping
        y -= 15
        if y < 50:  # new page
            p.showPage()
            y = 800

    p.showPage()
    p.save()
    buffer.seek(0)

    return send_file(
        buffer,
        mimetype="application/pdf",
        as_attachment=True,
        download_name="suppliers.pdf",
    )
@views.route("/admin/suppliers/import", methods=["POST"])
@login_required
def supplier_import():
    if not current_user.is_admin:
        abort(403)

    file = request.files.get("file")

    if not file or file.filename == "":
        flash("Please choose an Excel file to upload.", "error")
        return redirect(url_for("views.supplier_list"))

    filename = file.filename
    ext = filename.rsplit(".", 1)[-1].lower()
    allowed_exts = {"xlsx", "xlsm", "xltx", "xltm", "xls"}

    if ext not in allowed_exts:
        flash("Invalid file type. Allowed: .xls, .xlsx Excel files only.", "error")
        return redirect(url_for("views.supplier_list"))

    created = 0
    updated = 0
    skipped = 0

    try:
        rows = None

        if ext in {"xlsx", "xlsm", "xltx", "xltm"}:
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
        elif ext == "xls":
            data = file.read()
            book = xlrd.open_workbook(file_contents=data)
            sheet = book.sheet_by_index(0)
            rows = []
            for r in range(sheet.nrows):
                rows.append(sheet.row_values(r))

        if not rows or len(rows) < 2:
            flash("The uploaded file is empty or has no data rows.", "error")
            return redirect(url_for("views.supplier_list"))

        header_raw = rows[0]
        header = [
            (str(c).strip().lower() if c is not None else "")
            for c in header_raw
        ]

        def find_col(col_name):
            try:
                return header.index(col_name)
            except ValueError:
                return None

        name_idx = find_col("name")
        address_idx = find_col("address")
        email_idx = find_col("email")
        contact_idx = find_col("contact")

        if None in (name_idx, address_idx, email_idx, contact_idx):
            flash(
                "Header row must contain columns: Name, Address, Email, Contact.",
                "error",
            )
            return redirect(url_for("views.supplier_list"))

        def get_cell(row, i):
            if i is None or i >= len(row):
                return ""
            v = row[i]
            return "" if v is None else str(v).strip()

        for row in rows[1:]:
            if row is None:
                continue

            name = get_cell(row, name_idx)
            address = get_cell(row, address_idx)
            email = get_cell(row, email_idx)
            contact = get_cell(row, contact_idx)

            if not (name or address or email or contact):
                continue

            if not name or not email:
                skipped += 1
                continue

            existing = Supplier.query.filter_by(email=email).first()

            if existing:
                existing.name = name
                existing.address = address
                existing.contact = contact
                updated += 1
            else:
                s = Supplier(
                    name=name,
                    address=address,
                    email=email,
                    contact=contact,
                )
                db.session.add(s)
                created += 1

        db.session.commit()
        flash(
            f"Supplier import complete. Created: {created}, Updated: {updated}, Skipped: {skipped}.",
            "success",
        )

    except Exception as e:
        db.session.rollback()
        flash(f"Failed to import suppliers: {e}", "error")

    return redirect(url_for("views.supplier_list"))

@views.route("/admin/outgoing")
@login_required
def outgoing_list():
     ## this functions enable even user only
    # if not current_user.is_admin:
    #     abort(403)

    # pagination + search
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 10, type=int)
    search = (request.args.get("search") or "").strip()

    query = Outgoing.query.join(Product).join(Customer)

    if search:
        like = f"%{search}%"
        query = query.filter(
            or_(
                Product.name.ilike(like),
                Customer.name.ilike(like),
            )
        )

    pagination = query.order_by(Outgoing.date.desc(), Outgoing.id.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    outgoings = pagination.items

    # dropdown data for modal
    products = Product.query.order_by(Product.name).all()
    customers = Customer.query.order_by(Customer.name).all()

    return render_template(
        "admin_outgoing.html",
        user=current_user,
        outgoings=outgoings,
        pagination=pagination,
        products=products,
        customers=customers,
        search=search,
        per_page=per_page,
    )


@views.route("/admin/outgoing/new", methods=["POST"])
@login_required
def outgoing_create():
    if not current_user.is_admin:
        abort(403)

    product_id = request.form.get("product_id", type=int)
    customer_id = request.form.get("customer_id", type=int)
    qty_raw = (request.form.get("quantity") or "").strip()
    date_raw = (request.form.get("date") or "").strip()

    errors = []

    product = Product.query.get(product_id) if product_id else None
    customer = Customer.query.get(customer_id) if customer_id else None

    if not product:
        errors.append("Product is required.")
    if not customer:
        errors.append("Customer is required.")

    # quantity
    quantity = None
    try:
        quantity = int(qty_raw)
        if quantity <= 0:
            errors.append("Quantity must be a positive integer.")
    except Exception:
        errors.append("Quantity must be an integer.")

    # date
    if date_raw:
        try:
            date_val = datetime.strptime(date_raw, "%Y-%m-%d").date()
        except ValueError:
            errors.append("Invalid date format.")
            date_val = None
    else:
        date_val = datetime.utcnow().date()

    if errors:
        for e in errors:
            flash(e, "error")
        return redirect(url_for("views.outgoing_list"))

    record = Outgoing(
        product_id=product.id,
        customer_id=customer.id,
        quantity=quantity,
        date=date_val,
    )
    db.session.add(record)
    db.session.commit()

    flash("Outgoing product record created.", "success")
    return redirect(url_for("views.outgoing_list"))

@views.route("/admin/outgoing/<int:outgoing_id>/edit", methods=["POST"])
@login_required
def outgoing_edit(outgoing_id):
    if not current_user.is_admin:
        abort(403)

    record = Outgoing.query.get_or_404(outgoing_id)

    product_id = request.form.get("product_id", type=int)
    customer_id = request.form.get("customer_id", type=int)
    qty_raw = (request.form.get("quantity") or "").strip()
    date_raw = (request.form.get("date") or "").strip()

    errors = []

    product = Product.query.get(product_id) if product_id else None
    customer = Customer.query.get(customer_id) if customer_id else None

    if not product:
        errors.append("Product is required.")
    if not customer:
        errors.append("Customer is required.")

    try:
        quantity = int(qty_raw)
        if quantity <= 0:
            errors.append("Quantity must be a positive integer.")
    except Exception:
        errors.append("Quantity must be an integer.")
        quantity = None

    if date_raw:
        try:
            date_val = datetime.strptime(date_raw, "%Y-%m-%d").date()
        except ValueError:
            errors.append("Invalid date format.")
            date_val = None
    else:
        date_val = record.date

    if errors:
        for e in errors:
            flash(e, "error")
        return redirect(url_for("views.outgoing_list"))

    record.product_id = product.id
    record.customer_id = customer.id
    record.quantity = quantity
    record.date = date_val

    db.session.commit()
    flash("Outgoing product updated.", "success")
    return redirect(url_for("views.outgoing_list"))

@views.route("/admin/outgoing/<int:outgoing_id>/delete", methods=["POST"])
@login_required
def outgoing_delete(outgoing_id):
    if not current_user.is_admin:
        abort(403)

    record = Outgoing.query.get_or_404(outgoing_id)
    db.session.delete(record)
    db.session.commit()

    flash("Outgoing product deleted.", "success")
    return redirect(url_for("views.outgoing_list"))

@views.route("/admin/outgoing/export/excel")
@login_required
def outgoing_export_excel():
    if not current_user.is_admin:
        abort(403)

    search = (request.args.get("search") or "").strip()

    query = Outgoing.query.join(Product).join(Customer)
    if search:
        like = f"%{search}%"
        query = query.filter(
            or_(
                Product.name.ilike(like),
                Customer.name.ilike(like),
            )
        )

    rows = []
    for o in query.order_by(Outgoing.date.desc(), Outgoing.id.desc()).all():
        rows.append(
            {
                "ID": o.id,
                "Product": o.product.name if o.product else "",
                "Customer": o.customer.name if o.customer else "",
                "Quantity": o.quantity,
                "Date": o.date.isoformat() if o.date else "",
            }
        )

    df = pd.DataFrame(rows)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        df.to_excel(writer, index=False, sheet_name="Outgoing")
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="outgoing_products.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@views.route("/admin/outgoing/export/pdf")
@login_required
def outgoing_export_pdf():
    if not current_user.is_admin:
        abort(403)

    search = (request.args.get("search") or "").strip()

    query = Outgoing.query.join(Product).join(Customer)
    if search:
        like = f"%{search}%"
        query = query.filter(
            or_(
                Product.name.ilike(like),
                Customer.name.ilike(like),
            )
        )

    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)

    width, height = letter
    y = height - 50

    c.setFont("Helvetica-Bold", 14)
    c.drawString(30, y, "Outgoing Products List")
    y -= 30

    c.setFont("Helvetica-Bold", 10)
    c.drawString(30, y, "ID")
    c.drawString(60, y, "Product")
    c.drawString(200, y, "Customer")
    c.drawString(380, y, "Qty")
    c.drawString(420, y, "Date")
    y -= 20
    c.setFont("Helvetica", 9)

    for o in query.order_by(Outgoing.date.desc(), Outgoing.id.desc()).all():
        if y < 50:
            c.showPage()
            y = height - 50
            c.setFont("Helvetica", 9)
        c.drawString(30, y, str(o.id))
        c.drawString(60, y, (o.product.name if o.product else "")[:20])
        c.drawString(200, y, (o.customer.name if o.customer else "")[:20])
        c.drawString(380, y, str(o.quantity))
        c.drawString(420, y, o.date.isoformat() if o.date else "")
        y -= 15

    c.save()
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="outgoing_products.pdf",
        mimetype="application/pdf",
    )


@views.route("/admin/outgoing/<int:outgoing_id>/invoice")
@login_required
def outgoing_invoice(outgoing_id):
    if not current_user.is_admin:
        abort(403)

    o = Outgoing.query.get_or_404(outgoing_id)

    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    c.setFont("Helvetica-Bold", 16)
    c.drawString(200, height - 50, "INVOICE")

    y = height - 120
    c.setFont("Helvetica", 11)
    c.drawString(50, y, f"Invoice ID: {o.id}")
    y -= 20
    c.drawString(50, y, f"Date: {o.date.isoformat() if o.date else ''}")
    y -= 30

    c.drawString(50, y, "Customer:")
    y -= 20
    c.drawString(70, y, o.customer.name if o.customer else "")
    y -= 15
    if o.customer and o.customer.address:
        c.drawString(70, y, o.customer.address)
        y -= 15
    if o.customer and o.customer.email:
        c.drawString(70, y, f"Email: {o.customer.email}")
        y -= 15
    if o.customer and o.customer.contact:
        c.drawString(70, y, f"Contact: {o.customer.contact}")
        y -= 30

    # product table
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, y, "Product")
    c.drawString(300, y, "Quantity")
    c.drawString(380, y, "Unit Price")
    c.drawString(470, y, "Total")
    y -= 20
    c.setFont("Helvetica", 11)

    product_name = o.product.name if o.product else ""
    unit_price = o.product.price if (o.product and o.product.price is not None) else Decimal("0.00")
    total = unit_price * o.quantity

    c.drawString(50, y, product_name)
    c.drawString(300, y, str(o.quantity))
    c.drawString(380, y, f"{unit_price:.2f}")
    c.drawString(470, y, f"{total:.2f}")
    y -= 30

    c.setFont("Helvetica-Bold", 12)
    c.drawRightString(550, y, f"Grand Total: {total:.2f}")

    c.save()
    buffer.seek(0)

    filename = f"invoice_{o.id}.pdf"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/pdf",
    )


@views.route("/admin/purchases")
@login_required
def purchase_list():
    if not current_user.is_admin:
        abort(403)

    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 10, type=int)
    search = (request.args.get("search") or "").strip()

    query = Purchase.query.join(Product).join(Supplier)

    if search:
        like = f"%{search}%"
        query = query.filter(
            db.or_(
                Product.name.ilike(like),
                Supplier.name.ilike(like)
            )
        )

    pagination = query.order_by(Purchase.id.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    purchases = pagination.items

    products = Product.query.order_by(Product.name).all()
    suppliers = Supplier.query.order_by(Supplier.name).all()

    return render_template(
        "purchase_products.html",
        user=current_user,
        purchases=purchases,
        pagination=pagination,
        search=search,
        per_page=per_page,
        products=products,
        suppliers=suppliers,
    )


@views.route("/admin/purchases/new", methods=["POST"])
@login_required
def purchase_create():
    if not current_user.is_admin:
        abort(403)

    product_id = request.form.get("product_id")
    supplier_id = request.form.get("supplier_id")
    qty_raw = (request.form.get("quantity") or "").strip()
    date_raw = (request.form.get("date") or "").strip()

    errors = []

    product = Product.query.get(product_id) if product_id else None
    supplier = Supplier.query.get(supplier_id) if supplier_id else None

    if not product:
        errors.append("Product is required.")
    if not supplier:
        errors.append("Supplier is required.")

    quantity = None
    if not qty_raw:
        errors.append("Quantity is required.")
    else:
        try:
            quantity = int(qty_raw)
            if quantity <= 0:
                errors.append("Quantity must be greater than zero.")
        except ValueError:
            errors.append("Quantity must be an integer.")

    if date_raw:
        try:
            date_obj = datetime.strptime(date_raw, "%Y-%m-%d").date()
        except ValueError:
            errors.append("Invalid date format.")
    else:
        date_obj = datetime.utcnow().date()

    if errors:
        for e in errors:
            flash(e, "error")
        return redirect(url_for("views.purchase_list"))

    purchase = Purchase(
        product_id=product.id,
        supplier_id=supplier.id,
        quantity=quantity,
        date=date_obj,
    )
    db.session.add(purchase)

    # Optionally increase stock when purchasing
    product.quantity = (product.quantity or 0) + quantity

    db.session.commit()
    flash("Purchase record created.", "success")
    return redirect(url_for("views.purchase_list"))


@views.route("/admin/purchases/<int:purchase_id>/edit", methods=["POST"])
@login_required
def purchase_edit(purchase_id):
    if not current_user.is_admin:
        abort(403)

    purchase = Purchase.query.get_or_404(purchase_id)

    old_qty = purchase.quantity

    product_id = request.form.get("product_id")
    supplier_id = request.form.get("supplier_id")
    qty_raw = (request.form.get("quantity") or "").strip()
    date_raw = (request.form.get("date") or "").strip()

    errors = []

    product = Product.query.get(product_id) if product_id else None
    supplier = Supplier.query.get(supplier_id) if supplier_id else None

    if not product:
        errors.append("Product is required.")
    if not supplier:
        errors.append("Supplier is required.")

    quantity = None
    if not qty_raw:
        errors.append("Quantity is required.")
    else:
        try:
            quantity = int(qty_raw)
            if quantity <= 0:
                errors.append("Quantity must be greater than zero.")
        except ValueError:
            errors.append("Quantity must be an integer.")

    if date_raw:
        try:
            date_obj = datetime.strptime(date_raw, "%Y-%m-%d").date()
        except ValueError:
            errors.append("Invalid date format.")
    else:
        date_obj = datetime.utcnow().date()

    if errors:
        for e in errors:
            flash(e, "error")
        return redirect(url_for("views.purchase_list"))

    # adjust stock: remove old qty, add new qty
    if purchase.product_id == product.id:
        # same product
        product.quantity = (product.quantity or 0) - old_qty + quantity
    else:
        # different product
        old_product = Product.query.get(purchase.product_id)
        if old_product:
            old_product.quantity = (old_product.quantity or 0) - old_qty
        product.quantity = (product.quantity or 0) + quantity

    purchase.product_id = product.id
    purchase.supplier_id = supplier.id
    purchase.quantity = quantity
    purchase.date = date_obj

    db.session.commit()
    flash("Purchase updated.", "success")
    return redirect(url_for("views.purchase_list"))


@views.route("/admin/purchases/<int:purchase_id>/delete", methods=["POST"])
@login_required
def purchase_delete(purchase_id):
    if not current_user.is_admin:
        abort(403)

    purchase = Purchase.query.get_or_404(purchase_id)

    # adjust stock (remove purchased qty)
    product = purchase.product
    if product:
        product.quantity = (product.quantity or 0) - purchase.quantity

    db.session.delete(purchase)
    db.session.commit()

    flash("Purchase deleted.", "success")
    return redirect(url_for("views.purchase_list"))


@views.route("/admin/purchases/export/pdf")
@login_required
def purchases_export_pdf():
    if not current_user.is_admin:
        abort(403)

    purchases = Purchase.query.join(Product).join(Supplier).order_by(Purchase.id).all()

    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    y = height - 50
    c.setFont("Helvetica-Bold", 14)
    c.drawString(50, y, "Purchase Products List")
    y -= 30

    c.setFont("Helvetica", 10)
    for p in purchases:
        line = f"{p.id} - {p.product.name if p.product else ''} - {p.supplier.name if p.supplier else ''} - Qty: {p.quantity} - Date: {p.date}"
        c.drawString(50, y, line)
        y -= 15
        if y < 50:
            c.showPage()
            y = height - 50
            c.setFont("Helvetica", 10)

    c.save()
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name="purchases.pdf",
        mimetype="application/pdf",
    )


@views.route("/admin/purchases/export/excel")
@login_required
def purchases_export_excel():
    if not current_user.is_admin:
        abort(403)

    purchases = Purchase.query.join(Product).join(Supplier).order_by(Purchase.id).all()

    rows = []
    for p in purchases:
        rows.append({
            "ID": p.id,
            "Product": p.product.name if p.product else "",
            "Supplier": p.supplier.name if p.supplier else "",
            "Quantity": p.quantity,
            "Date": p.date.isoformat() if p.date else "",
        })

    df = pd.DataFrame(rows)

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Purchases")

    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name="purchases.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@views.route("/admin/purchases/<int:purchase_id>/invoice/pdf")
@login_required
def purchase_invoice_pdf(purchase_id):
    if not current_user.is_admin:
        abort(403)

    p = Purchase.query.get_or_404(purchase_id)

    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    c.setFont("Helvetica-Bold", 16)
    c.drawString(50, height - 50, "Purchase Invoice")

    c.setFont("Helvetica", 11)
    y = height - 100
    c.drawString(50, y, f"Invoice ID: {p.id}")
    y -= 20
    c.drawString(50, y, f"Date: {p.date.isoformat() if p.date else ''}")
    y -= 20
    c.drawString(50, y, f"Product: {p.product.name if p.product else ''}")
    y -= 20
    c.drawString(50, y, f"Supplier: {p.supplier.name if p.supplier else ''}")
    y -= 20
    c.drawString(50, y, f"Quantity: {p.quantity}")

    c.showPage()
    c.save()
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"purchase_invoice_{p.id}.pdf",
        mimetype="application/pdf",
    )


@views.route("/admin/users")
@login_required
@roles_required('admin')
def system_users_list():
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 10, type=int)
    search = (request.args.get("search") or "").strip()

    query = User.query

    if search:
        like = f"%{search}%"
        query = query.filter(
            db.or_(
                User.first_name.ilike(like),
                User.email.ilike(like)
            )
        )

    pagination = query.order_by(User.id).paginate(
        page=page, per_page=per_page, error_out=False
    )
    users = pagination.items

    return render_template(
        "system_users.html",
        user=current_user,
        users=users,
        pagination=pagination,
        per_page=per_page,
        search=search,
    )



@views.route("/admin/users/new", methods=["POST"])
@login_required
@roles_required('admin')
def system_user_create():
    name = (request.form.get("name") or "").strip()
    email = (request.form.get("email") or "").strip().lower()
    role = (request.form.get("role") or "user").strip()
    password = request.form.get("password") or ""
    confirm = request.form.get("confirm_password") or ""

    errors = []

    if not name:
        errors.append("Name is required.")
    if not email:
        errors.append("Email is required.")
    else:
        existing = User.query.filter_by(email=email).first()
        if existing:
            errors.append("Email already exists.")

    if role not in ("admin", "user"):
        errors.append("Invalid role selected.")

    if not password:
        errors.append("Password is required.")
    elif len(password) < 6:
        errors.append("Password must be at least 6 characters.")

    if password != confirm:
        errors.append("Password and confirm password do not match.")

    if errors:
        for e in errors:
            flash(e, "error")
        return redirect(url_for("views.system_users_list"))

    hashed = generate_password_hash(password, method="scrypt")

    # Create user using `role` column if model supports it; fall back to is_admin
    kwargs = dict(first_name=name, email=email, password=hashed)
    if hasattr(User, "role"):
        kwargs["role"] = role
    else:
        kwargs["is_admin"] = (role == "admin")

    new_user = User(**kwargs)
    db.session.add(new_user)
    db.session.commit()

    flash("User registered successfully.", "success")
    return redirect(url_for("views.system_users_list"))


@views.route("/admin/users/<int:user_id>/edit", methods=["POST"])
@login_required
@roles_required('admin')
def system_user_edit(user_id):
    user = User.query.get_or_404(user_id)

    name = (request.form.get("name") or "").strip()
    email = (request.form.get("email") or "").strip().lower()
    role = (request.form.get("role") or "user").strip()
    password = request.form.get("password") or ""
    confirm = request.form.get("confirm_password") or ""

    errors = []

    if not name:
        errors.append("Name is required.")

    if not email:
        errors.append("Email is required.")
    else:
        existing = User.query.filter_by(email=email).first()
        if existing and existing.id != user.id:
            errors.append("Email already exists for another user.")

    if role not in ("admin", "user"):
        errors.append("Invalid role selected.")

    if password:
        if len(password) < 6:
            errors.append("New password must be at least 6 characters.")
        if password != confirm:
            errors.append("New password and confirmation do not match.")

    if errors:
        for e in errors:
            flash(e, "error")
        return redirect(url_for("views.system_users_list"))

    # Danger checks: prevent demoting the last admin and prevent self-demotion
    # Determine user's current admin status and count other admins
    def is_user_admin(u):
        return (getattr(u, "role", None) == "admin") or getattr(u, "is_admin", False)

    # If the edit will demote this user from admin -> non-admin, ensure there is at least one other admin
    will_be_admin = (role == "admin")
    currently_admin = is_user_admin(user)

    if currently_admin and not will_be_admin:
        # count other admins besides this user
        if hasattr(User, "role"):
            other_admins = User.query.filter(User.role == 'admin', User.id != user.id).count()
        else:
            other_admins = User.query.filter_by(is_admin=True).filter(User.id != user.id).count()

        if other_admins == 0:
            flash("Cannot demote this user — there must be at least one admin account.", "error")
            return redirect(url_for("views.system_users_list"))

        # also prevent self-demotion
        if user.id == current_user.id:
            flash("You cannot remove your own admin role. Ask another admin to change your role.", "error")
            return redirect(url_for("views.system_users_list"))

    # Save changes
    user.first_name = name
    user.email = email
    if hasattr(User, "role"):
        user.role = role
    else:
        user.is_admin = (role == "admin")

    if password:
        user.password = generate_password_hash(password, method="scrypt")

    db.session.commit()
    flash("User updated successfully.", "success")
    return redirect(url_for("views.system_users_list"))

@views.route("/admin/users/<int:user_id>/delete", methods=["POST"])
@login_required
@roles_required('admin')
def system_user_delete(user_id):
    user_obj = User.query.get_or_404(user_id)

    # Don’t let a user delete themselves
    if user_obj.id == current_user.id:
        flash("You cannot delete your own account.", "error")
        return redirect(url_for("views.system_users_list"))

    # Don’t delete the last admin
    def is_admin(u):
        return (getattr(u, "role", None) == "admin") or getattr(u, "is_admin", False)

    if is_admin(user_obj):
        if hasattr(User, "role"):
            admins_count = User.query.filter(User.role == 'admin').count()
        else:
            admins_count = User.query.filter_by(is_admin=True).count()

        if admins_count <= 1:
            flash("You cannot delete the last admin user.", "error")
            return redirect(url_for("views.system_users_list"))

    db.session.delete(user_obj)
    db.session.commit()
    flash("User deleted.", "success")
    return redirect(url_for("views.system_users_list"))




